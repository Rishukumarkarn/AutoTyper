from tkinter import *
from tkinter import ttk,filedialog,messagebox
import openpyxl
from time import sleep
import keyboard
from PIL import ImageTk, Image

class Auto_typer:

    def __init__(self,root):
        self.root=root
        self.root.title('Auto Robot')
        self.root.geometry("220x320")
        self.root.resizable(0, 0)
        self.root.tk.call('wm', 'iconphoto', Tk._w, ImageTk.PhotoImage(Image.open('./res/tt.png')))


        frst_btn = Button(self.root, text='Load Excel File[+]',command=self.load_excel,bg='#034efc', font=("time new roman", 10, "bold")).place(x=10, y=5,width=200,height=25)

        # All Variable............................................................!
        self.var_excel_file=StringVar()
        self.var_laod_column=StringVar()
        self.var_list_value=IntVar()
        self.var_tm_dely=StringVar()
        self.var_spd_value=StringVar()



        # Design start..........................
       # 1st Frame with Scrollbar......................................................................!
        frst_frame=Frame(self.root,relief=RIDGE,bd=2)
        frst_frame.place(x=10,y=35,width=200,height=210)

        scrollbar=Scrollbar(frst_frame)
        scrollbar.pack(side=RIGHT,fill=Y)
        self.d_list=Listbox(frst_frame,yscrollcommand=scrollbar.set)
        self.d_list.pack(side=LEFT,fill=BOTH,expand=True)
        scrollbar.config(command=self.d_list.yview)

        # 2nd Frame.........................................................................!
        snd_frame = Frame(self.root, relief=RIDGE,bg='black')
        snd_frame.place(x=10, y=255, width=200, height=78)

        snd_f_btn=Button(snd_frame,text='Transfer Text',command=self.select_list_value,bg='#5603fc', font=("time new roman",7, "bold")).place(width=80,height=30)

        lbl_f_snd=Label(snd_frame,text='Time Delay :', bg="#000000",fg="dark goldenrod", font=("time new roman", 8, "bold")).place(x=85,y=5)

        combo_f_snd=ttk.Combobox(snd_frame,width=3,font=("time new roman", 10, "bold"), state="readonly",textvariable=self.var_tm_dely)
        combo_f_snd['value']=('2s','5s','10s','15s')
        combo_f_snd.current(0)
        combo_f_snd.place(x=170,y=5,width=30)

        lbl_s_snd = Label(snd_frame, text='Text Speed (WPM) :', bg="#000000", fg="dark goldenrod",font=("time new roman", 7, "bold")).place( y=40)

        combo_s_snd = ttk.Combobox(snd_frame, width=3, font=("time new roman", 10, "bold"), state="readonly",textvariable=self.var_spd_value)
        combo_s_snd['value'] = ('0.1', '0.2', '0.3', '0.4')
        combo_s_snd.current(0)
        combo_s_snd.place(x=105, y=40, width=39)

        snd_s_btn = Button(snd_frame, text='Reload',command=self.refresh,bg='#fc2003', font=("time new roman", 8, "bold")).place(width=50, height=23,x=152, y=40)

        # Design Complite..........................


    # Funcation start.....................................


    def load_excel(self):
        try:
            file=filedialog.askopenfilename()
            if file!=None:
                self.var_excel_file.set(str(file))
                path=self.var_excel_file.get()
                df=openpyxl.load_workbook(path)
                read=df.active
                m=read.max_row
                for i in range(2,m + 1):
                    reads=read.cell(column=2,row=i)
                    load_file=(reads.value)
                    self.var_laod_column.set(load_file)
                    self.d_list.insert(END,self.var_laod_column.get())
        except:
            pass


    def select_list_value(self):
        try:
            n=self.d_list.curselection()
            for i in n:
                if i==0:
                    m=i+2
                    print(m)
                    self.var_list_value.set(m)
                else:
                    m=i+2
                    # print(m)
                    self.var_list_value.set(m)
            self.type_select_value()
        except:
            messagebox.showwarning("Empty", "Please select a excel List!")


    def type_select_value(self):
        r = int(self.var_list_value.get())
        t=int(self.var_tm_dely.get().split('s',1)[0])
        sleep(t)
        df=openpyxl.load_workbook(self.var_excel_file.get())
        read=df.active
        m=read.max_column
        for i in range(2, m + 1):
            sd=float(self.var_spd_value.get())
            sleep(sd)
            reads=read.cell(row=r,column=i)
            load_file=(reads.value)
            keyboard.write(load_file)
            keyboard.press_and_release('tab')

    def refresh(self):
        self.__init__(root)

root=Tk()
aut=Auto_typer(root)
root.configure(bg='black')
root.mainloop()